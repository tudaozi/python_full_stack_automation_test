#!/usr/bin/env python
# coding=UTF-8
"""
@Author: STAURL.COM
@Contact: admin@staurl.com
@Project: python_full_stack_automation_test
@File: test_cw0625_excel_package_1.py
@Time: 2019-06-30 11:44
@Desc: S
"""
from openpyxl import load_workbook  # 导入openpyxl第三方库


class HandleExcel:  # 创建一个Excel处理类
    def __init__(self, filename, sheetname=None):  # 初始化类，并设置两个参数
        self.filename = filename
        self.sheetname = sheetname

    def load_excel(self):
        wb = load_workbook(self.filename)  # 加载Excel文件
        if self.sheetname is None:  # 判断是否有指定表单，如果没有使用打开Excel文件的激活表单，如果有就是用指定名称的表单
            ws = wb.active
        else:
            ws = wb[self.sheetname]
        return wb, ws

    def get_case(self):  # 获取行-测试用例
        cases_list = []  # 创建一个空的用例列表
        cases_title_entity = tuple(self.load_excel()[1].iter_rows(max_row=1, values_only=True))
        cases_title = cases_title_entity[0]
        for cases_entity in tuple(self.load_excel()[1].iter_rows(min_row=2, values_only=True)):
            cases_list.append(dict(zip(cases_title, cases_entity)))
        return cases_list  # 将用例列表返回给当前函数

    def write_result(self, row, actual, result):
        if isinstance(row, int) and (2 <= row <= self.load_excel()[1].max_row):
            # self.load_excel()[1].cell(row=row, column=6, value=actual)
            # self.load_excel()[1].cell(row=row, column=7, value=result)
            self.load_excel()[1].cell(3, 6).value == actual
            self.load_excel()[1].cell(3, 7).value = result
            print(self.load_excel()[1].cell(3, 7).value)
            print(self.load_excel()[1])
            self.load_excel()[0].save(self.filename)



if __name__ == '__main__':  # 在当前路径下才可以运行
    file_name = 'demo.xlsx'  # 指定文件
    my_excel = HandleExcel(file_name, 'add')  # 实例化my_excel这个对象
    print(my_excel.get_case())  # 将对象调用方法后的结果打印出来
    my_excel.write_result(3, 1, 2)
