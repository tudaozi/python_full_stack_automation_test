#!/usr/bin/env python
# coding=UTF-8
"""
@Author: STAURL.COM
@Contact: admin@staurl.com
@Project: python_full_stack_automation_test
@File: cw0615.py
@Time: 2019-06-23 21:57
@Desc: S
"""

from openpyxl import load_workbook, Workbook


# import os


# 1、总结类和对象的所有知识点。思维导图形式或者文字。

# 2、在之前定义的手机类下面定义智能手机类和苹果手机类。智能手机打电话具有录音功能。 苹果手机打电话不仅具有录音功能，还有 facetime 功能
# class MobilePhone:
#     screen = 'Touch'
#     camera = 'Megapixel'
#
#     def __init__(self, brand, color, model='OrdinaryPhone', system_type='Android', cellphone_number=None, caller=None,
#                  answer=None):
#         self.brand = brand
#         self.color = color
#         self.model = model
#         self.system_type = system_type
#         self.cellphone_number = cellphone_number
#         self.caller = caller
#         self.answer = answer
#         print('品牌：{}，颜色：{}，类型：{}，系统：{}'.format(brand, color, model, system_type))
#
#     def card(self):
#         return self.cellphone_number
#
#     def call(self):
#         print('{}使用{}号码打电话给{}'.format(self.caller, self.card(), self.answer))
#
#     @staticmethod
#     def gap():
#         print('\\\\', '*' * 69)
#
#
# class SmartPhone(MobilePhone):
#     def __init__(self, brand, color, record_content, model='OrdinaryPhone', system_type='Android',
#                  cellphone_number=None, caller=None,
#                  answer=None):
#         super(SmartPhone, self).__init__(brand, color, model='OrdinaryPhone', system_type='Android')
#         self.brand = brand
#         self.color = color
#         self.record_content = record_content
#         self.model = model
#         self.system_type = system_type
#         self.cellphone_number = cellphone_number
#         self.caller = caller
#         self.answer = answer
#
#     def call(self, record=False):
#         super(SmartPhone, self).call()
#         if record:
#             print('{}使用{}号码打电话给{}说{}'.format(self.caller, self.card(), self.answer, self.record()))
#
#     def record(self):
#         return self.record_content
#
#
# class ApplePhone(SmartPhone):
#     def call(self, record=False, facetime=False):
#         super(ApplePhone, self).call(record)
#         if facetime:
#             print('{}使用{}号码打电话给{}{}'.format(self.caller, self.card(), self.answer, self.facetime()))
#
#     def facetime(self):
#         return '正在FaceTime通话'
#
#
# my_phone1 = MobilePhone('树树', '红色', cellphone_number=13888888888, caller='刀刀', answer='树树')
# my_phone1.gap()
# my_phone2 = SmartPhone('树树', '红色', '我想你', cellphone_number=13888888888, caller='刀刀', answer='树树')
# my_phone2.call(record=True)
# my_phone1.gap()
# my_phone3 = ApplePhone('树树', '红色', '我想你', cellphone_number=13888888888, caller='刀刀', answer='树树')
# my_phone3.call(record=False, facetime=True)
# my_phone1.gap()
# my_phone3.call(record=True, facetime=False)
# my_phone1.gap()
# my_phone3.call(record=True, facetime=True)
# my_phone1.gap()
# my_phone3.call(record=False, facetime=False)


# 3、定义一个 ExcelManual 类。具有获取 sheet 表单， 读取单元格 和 修改单元格功能。
class ExcelManual:
    def __init__(self, path):
        self.path = path
        self.wb = None

    def create_excle(self):
        wb = Workbook()

        # grab the active worksheet
        ws = wb.active

        # Data can be assigned directly to cells
        ws['A1'] = 42

        # Rows can also be appended
        ws.append([1, 2, 3])

        # Save the file
        wb.save("sample.xlsx")

    def open_excle(self):
        load_file = load_workbook(self.path)
        print(load_file)

    def get_space(self):
        self.wb = Workbook()
        ws = self.wb.active
        ws.append([1, 2, 3])

    def save_excle(self):
        self.wb.save("sample.xlsx")
        pass
    #
    # def close_excle(self):
    #     pass
    #
    # def get_the_form(self):
    #     pass
    #
    # def read_cell(self):
    #     pass
    #
    # def modify_cell(self):
    #     pass
    #
    # pass


my_excle = ExcelManual('demo.xlsx')
my_excle.save_excle()
# print(os.getcwd())
