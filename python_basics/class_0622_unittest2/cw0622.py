#!/usr/bin/env python
# coding=UTF-8
"""
@Author: STAURL.COM
@Contact: admin@staurl.com
@Project: python_full_stack_automation_test
@File: cw0622.py
@Time: 2019/6/26 9:03
@Desc: S
"""
from openpyxl import load_workbook, Workbook


class OneExcel:
    def __init__(self, path):
        self.path = path

    def open_excel(self):
        wb = Workbook()
        ws = wb.active
        ws.append(range(10))
        wd = tuple(ws.iter_rows(values_only=True))
        wd1 = ''
        for i in wd[0]:
            wd1 += str(eval('i'))
        wb.save(self.path)
        print(wd1)


if __name__ == '__main__':
    my_excel = OneExcel('demo.xlsx')
    my_excel.open_excel()
